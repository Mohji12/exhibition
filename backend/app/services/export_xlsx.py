from __future__ import annotations

import io
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font

from app.schemas import AdminOverview, Lead


def build_leads_workbook(
    leads: list[Lead],
    overview: AdminOverview,
    summary_markdown: str,
    image_flags: dict[str, bool] | None = None,
) -> bytes:
    image_flags = image_flags or {}
    wb = Workbook()

    ws_leads = wb.active
    ws_leads.title = "Leads"
    headers = [
        "id",
        "name",
        "company",
        "designation",
        "mobile",
        "email",
        "city",
        "priority",
        "interests",
        "synced",
        "captured_at",
        "capture_source",
        "capturer_name",
        "capturer_email",
        "summary",
        "ocr_quality",
        "ai_issues",
        "has_card_image",
    ]
    ws_leads.append(headers)
    for cell in ws_leads[1]:
        cell.font = Font(bold=True)

    for lead in leads:
        meta = lead.capture_meta
        ws_leads.append(
            [
                lead.id,
                lead.name,
                lead.company,
                lead.designation,
                lead.mobile,
                str(lead.email),
                lead.city,
                lead.priority,
                "|".join(lead.interests),
                1 if lead.synced else 0,
                lead.captured_at,
                lead.capture_source or "",
                lead.capturer_name or "",
                str(lead.capturer_email or ""),
                lead.summary,
                (meta.ocr_quality if meta else "") or "",
                "; ".join(meta.ai_issues) if meta and meta.ai_issues else "",
                1 if image_flags.get(lead.id) else 0,
            ]
        )

    ws_stats = wb.create_sheet("Stats")
    ws_stats.append(["Metric", "Value"])
    for cell in ws_stats[1]:
        cell.font = Font(bold=True)
    rows: Iterable[tuple[str, object]] = [
        ("Total leads", overview.leads),
        ("Hot", overview.hot_leads),
        ("Warm", overview.warm_leads),
        ("Cold", overview.cold_leads),
        ("Synced", overview.synced_leads),
        ("Unsynced", overview.unsynced_leads),
        ("Pending follow-ups", overview.pending_follow_ups),
        ("Staff active", overview.staff_active),
        ("Source QR", overview.by_source.qr),
        ("Source card", overview.by_source.card),
        ("Source manual", overview.by_source.manual),
        ("Source unknown", overview.by_source.unknown),
    ]
    for metric, value in rows:
        ws_stats.append([metric, value])
    ws_stats.append([])
    ws_stats.append(["Top interest", "Count"])
    for interest in overview.top_interests:
        ws_stats.append([interest.name, interest.count])

    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Booth report"
    ws_summary["A1"].font = Font(bold=True, size=14)
    # Split markdown into rows for readability in Excel
    for idx, line in enumerate((summary_markdown or "").splitlines() or [""], start=3):
        ws_summary.cell(row=idx, column=1, value=line)
    ws_summary.column_dimensions["A"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
